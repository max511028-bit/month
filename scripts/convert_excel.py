import json
import re
import sys
from pathlib import Path

import openpyxl


SOURCE = Path(sys.argv[1] if len(sys.argv) > 1 else "../source.xlsx")
OUTPUT_DIR = Path(sys.argv[2] if len(sys.argv) > 2 else "data")

DIVISION_LEADERS = {
    "Урал": "Николай Шмаков",
    "5 Дивизион": "Юлия Кузьменко",
    "5 дивизион": "Юлия Кузьменко",
    "Южный дивизион": "Иразитдин Халатов",
}

SHEET_MONTHS = {
    "МАРТ": {"id": "2026-03", "label": "Март 2026", "short": "Март"},
    "АПРЕЛЬ": {"id": "2026-04", "label": "Апрель 2026", "short": "Апрель"},
}

METRIC_MAP = {
    "request": 3,
    "avgOutput": 4,
    "currentStaff": 5,
    "productivity": 6,
    "penalties": 7,
    "penaltyShare": 8,
    "revenueVat": 9,
    "requestCloseRate": 10,
    "secretCheck": 11,
    "realizationPayroll": 12,
    "recruitingPayroll": 15,
    "invited": 16,
    "invitedToResponseRate": 17,
    "registered": 18,
    "registeredToInvitedRate": 19,
    "warehouseReached": 20,
    "warehouseToRegisteredRate": 21,
    "firstShift": 22,
    "firstShiftToWarehouseRate": 23,
    "tenShifts": 24,
    "tenShiftsToFirstShiftRate": 25,
    "responseToWarehouseRate": 28,
    "marketingProjectName": 29,
    "marketingPayroll": 30,
    "marketingBudget": 31,
    "responses": 32,
    "targetLeads": 33,
    "targetLeadRate": 34,
    "responseCost": 37,
    "targetLeadCost": 38,
}

TEXT_MAP = {
    "realizationDrivers": 13,
    "realizationBlockers": 14,
    "recruitingDrivers": 26,
    "recruitingBlockers": 27,
    "marketingDrivers": 35,
    "marketingBlockers": 36,
    "agreements": 39,
}

METRIC_LABELS = {
    "request": "Заявка от клиента",
    "avgOutput": "Суточный выход",
    "currentStaff": "Текущий штат",
    "productivity": "Производительность",
    "penalties": "Штрафные санкции",
    "penaltyShare": "% штрафов от выручки",
    "revenueVat": "Выручка с НДС",
    "requestCloseRate": "% закрытия заявки",
    "secretCheck": "Проверка СУПР",
    "realizationPayroll": "ФОТ реализации",
    "recruitingPayroll": "ФОТ подбора",
    "invited": "Приглашенные",
    "registered": "Оформленные",
    "warehouseReached": "Дошедшие до склада",
    "firstShift": "Вышедшие на 1 смену",
    "tenShifts": "Вышедшие на 10 смен",
    "marketingPayroll": "ФОТ маркетинга",
    "marketingBudget": "Бюджет маркетинга",
    "responses": "Отклики",
    "targetLeads": "Целевые лиды",
    "targetLeadRate": "% целевых лидов",
    "responseCost": "Стоимость отклика",
    "targetLeadCost": "Стоимость целевого",
}


def slug(value):
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-zа-яё0-9]+", "-", text, flags=re.I).strip("-")
    return text or "empty"


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if text in {"", "-", "—"}:
            return None
        return re.sub(r"\s+", " ", text)
    return value


def number(value):
    value = clean(value)
    if isinstance(value, (int, float)):
        return float(value)
    return None


def text(value):
    value = clean(value)
    return value if isinstance(value, str) else None


def merged_value(ws, row, col):
    cell = ws.cell(row, col)
    if clean(cell.value) is not None:
        return clean(cell.value)
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return clean(ws.cell(rng.min_row, rng.min_col).value)
    return None


def action_category(body):
    source = (body or "").lower()
    rules = [
        ("ЛМК / документы", ["лмк", "справк", "документ", "оформлен"]),
        ("Ставка / тариф", ["ставк", "тариф", "зп", "зарплат"]),
        ("Конкуренты", ["конкур", "ка ", "кредо", "пролог"]),
        ("Заявка клиента", ["заявк", "потребност", "объем", "обьем"]),
        ("Жилье / вахта", ["вахт", "жиль", "хостел"]),
        ("Маркетинг", ["реклам", "лид", "отклик", "трафик"]),
        ("Качество персонала", ["качеств", "рп", "персонал", "сотрудник"]),
        ("Склад / условия", ["склад", "услов", "график"]),
        ("Руководитель / СУПР", ["супр", "бригад", "руковод"]),
        ("Штрафы", ["штраф", "шс", "инвентаризац"]),
    ]
    for category, needles in rules:
        if any(needle in source for needle in needles):
            return category
    return "Другое"


def default_norms():
    company = {
        "requestCloseRate": {"target": 0.95, "direction": "gte", "unit": "%", "weight": 16},
        "secretCheck": {"target": 80, "direction": "gte", "unit": "score", "weight": 8},
        "penaltyShare": {"target": 0.03, "direction": "lte", "unit": "%", "weight": 10},
        "targetLeadRate": {"target": 0.22, "direction": "gte", "unit": "%", "weight": 9},
        "responseCost": {"target": 200, "direction": "lte", "unit": "rub", "weight": 8},
        "targetLeadCost": {"target": 900, "direction": "lte", "unit": "rub", "weight": 9},
        "registeredToInvitedRate": {"target": 0.55, "direction": "gte", "unit": "%", "weight": 8},
        "warehouseToRegisteredRate": {"target": 0.65, "direction": "gte", "unit": "%", "weight": 8},
        "firstShiftToWarehouseRate": {"target": 0.75, "direction": "gte", "unit": "%", "weight": 8},
        "tenShiftsToFirstShiftRate": {"target": 0.45, "direction": "gte", "unit": "%", "weight": 8},
        "revenueVat": {"target": None, "direction": "gte", "unit": "rub", "weight": 8},
        "penalties": {"target": None, "direction": "lte", "unit": "rub", "weight": 4},
        "productivity": {"target": None, "direction": "gte", "unit": "number", "weight": 4},
    }
    return {
        "scopeOrder": ["project", "division", "company"],
        "company": company,
        "division": {},
        "project": {},
        "month": {},
        "notes": [
            "Нормативы изменяемые: компания, дивизион, проект и месяц.",
            "Если у проекта нет своего норматива, используется норматив дивизиона, затем компании.",
        ],
    }


def parse_sheet(wb, sheet_name):
    ws = wb[sheet_name]
    month = SHEET_MONTHS[sheet_name]
    report = {
        "id": month["id"],
        "label": month["label"],
        "sourceSheet": sheet_name,
        "projects": [],
        "totals": [],
        "summary": None,
    }
    current_division = None
    final_prefix = "Итоги"

    for row in range(3, ws.max_row + 1):
        raw_group = merged_value(ws, row, 1)
        project_name = text(ws.cell(row, 2).value)

        if isinstance(raw_group, str) and raw_group.startswith(final_prefix):
            report["summary"] = parse_metrics(ws, row)
            report["summary"]["label"] = raw_group
            break

        if isinstance(raw_group, str) and raw_group.startswith("Итог"):
            total = parse_metrics(ws, row)
            total["label"] = raw_group
            total["division"] = current_division
            report["totals"].append(total)
            continue

        if raw_group:
            current_division = raw_group

        if not project_name:
            continue

        division = current_division or raw_group or "Без дивизиона"
        metrics = parse_metrics(ws, row)
        comments = {key: text(ws.cell(row, col).value) for key, col in TEXT_MAP.items()}
        project = {
            "id": slug(project_name),
            "name": project_name,
            "division": division,
            "divisionLeader": DIVISION_LEADERS.get(division, ""),
            "monthId": month["id"],
            "metrics": metrics,
            "comments": comments,
        }
        report["projects"].append(project)
    return report


def parse_metrics(ws, row):
    data = {}
    for key, col in METRIC_MAP.items():
        if key == "marketingProjectName":
            data[key] = text(ws.cell(row, col).value)
        else:
            data[key] = number(ws.cell(row, col).value)
    return data


def build_actions(reports):
    actions = []
    status_by_type = {
        "driver": "в работе",
        "blocker": "новое",
        "agreement": "новое",
    }
    fields = [
        ("realizationDrivers", "Реализация", "driver", "Драйвер роста"),
        ("realizationBlockers", "Реализация", "blocker", "Блокер"),
        ("recruitingDrivers", "Подбор", "driver", "Драйвер роста"),
        ("recruitingBlockers", "Подбор", "blocker", "Блокер"),
        ("marketingDrivers", "Маркетинг", "driver", "Драйвер роста"),
        ("marketingBlockers", "Маркетинг", "blocker", "Блокер"),
        ("agreements", "Управление", "agreement", "Договоренность"),
    ]
    for report in reports:
        for project in report["projects"]:
            for key, department, kind, title in fields:
                body = project["comments"].get(key)
                if not body:
                    continue
                actions.append({
                    "id": f"{report['id']}-{project['id']}-{key}",
                    "monthId": report["id"],
                    "projectId": project["id"],
                    "projectName": project["name"],
                    "division": project["division"],
                    "department": department,
                    "kind": kind,
                    "title": title,
                    "text": body,
                    "category": action_category(body),
                    "status": status_by_type[kind],
                    "owner": project["divisionLeader"],
                    "dueDate": None,
                })
    return actions


def main():
    wb = openpyxl.load_workbook(SOURCE, data_only=True)
    reports = [parse_sheet(wb, sheet) for sheet in SHEET_MONTHS]

    projects = {}
    for report in reports:
        for item in report["projects"]:
            saved = projects.setdefault(item["id"], {
                "id": item["id"],
                "name": item["name"],
                "divisions": [],
                "mediaNames": [],
            })
            if item["division"] not in saved["divisions"]:
                saved["divisions"].append(item["division"])
            media = item["metrics"].get("marketingProjectName")
            if media and media not in saved["mediaNames"]:
                saved["mediaNames"].append(media)

    payload = {
        "generatedAt": "2026-05-13",
        "source": "Google Sheets export",
        "metricLabels": METRIC_LABELS,
        "divisionLeaders": DIVISION_LEADERS,
        "projects": sorted(projects.values(), key=lambda x: x["name"]),
        "reports": reports,
    }

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUTPUT_DIR / "monthly-reports.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (OUTPUT_DIR / "norms.json").write_text(
        json.dumps(default_norms(), ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (OUTPUT_DIR / "actions.json").write_text(
        json.dumps({"actions": build_actions(reports)}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
